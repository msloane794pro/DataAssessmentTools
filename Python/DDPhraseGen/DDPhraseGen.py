### Data Dictionary Phrase Generator
# Tool to create Friendly Names and Descriptions of data elements defined in a data dictionary.
# Utilizes the OpenAiDDPhraseGen class to to access and use AI models and natural language processing.
# Copyright 2024 - Michael Sloane - Agree Technologies

#
# Import needed libraries
import os
import sys
import time
import argparse
from datetime import datetime
import pandas as pd
import openpyxl
from io import StringIO
import csv

from OpenAiDDPhraseGen import PhraseType as pt
from OpenAiDDPhraseGen import OpenAiDDPhraseGen

# Define constants
dateLastUpdated = '2024.05.20 03:26:34'
maxRetries = 3


def run_description_program(dbDescr, tables, columnData, output_file, phraseType):
    print(f'Running Description program with output: {output_file}, type: {phraseType}')
    print(f'Input data contains {columnData.shape[0]} rows.')
    pgen = OpenAiDDPhraseGen()
    modelList = pgen.getModelNames()
    iterationCount = 1
    failedTables = tables['TableName'].tolist()
    column_names = ['TableName', 'ColumnName', 'Friendly Name', 'Description']
    sessionData = pd.DataFrame(columns=column_names)
    tableData = pd.DataFrame(tables)

    tablesDf = tableData

    while iterationCount <= maxRetries:
        for i in range(len(modelList)):
            print(f'-->\nRunning iteration {iterationCount} on AI Model `{modelList[i]}` on {len(failedTables)} tables.')

            currentSessionData, currentFailedTables = run_descr_session(pgen, dbDescr, tablesDf, columnData, phraseType, output_file, 0)
            
            if len(currentFailedTables) < len(failedTables):
                sessionData = pd.concat([sessionData, currentSessionData], ignore_index=True)
                failedTables = currentFailedTables
                tablesDf = pd.DataFrame(failedTables, columns=['TableName'])

            if len(currentFailedTables) == 0:
                break

        if len(failedTables) == 0:
            break

        iterationCount += 1    


    tableData['AlreadyInDataHub'] = 'N'
    newOrder = ['TableName', 'AlreadyInDataHub', 'Description']
    tableData = tableData[newOrder]

    sessionData['IncludeInView'] = 'Y'
    newOrder = ['TableName', 'ColumnName', 'IncludeInView', 'Friendly Name', 'Description']
    sessionData = sessionData[newOrder]
    sessionDf = sessionData.sort_values(by=['TableName', 'ColumnName'])
    if len(failedTables) > 0:
        print(f'Unable to create Descriptions for the following Tables: {failedTables}')

    completedNumRows = sessionDf.shape[0]
    print(f'Completed session data contains {completedNumRows} rows.')
    descriptions_generated = sessionDf.query("`Description`!= ''")
    descriptionCount = descriptions_generated.shape[0]
    print(f'Descriptions generated: {descriptionCount} ({format(100 * descriptionCount/completedNumRows, ".1f")}%).')
    
    tableData.to_excel(output_file, sheet_name='Table Descriptions', index=False)
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        sessionDf.to_excel(writer, sheet_name='Column Descriptions', index=False)

    print(f'Complete generated data saved to: {output_file}')

    return


def dataframe_to_dict(df):
    return df.to_dict('list')


def run_friendlyName_program(dbDescr, tables, glossary, output_file, phraseType):
    print(f'Running Friendly program with output: {output_file}, type: {phraseType}')
    print(f'Input data contains {glossary.shape[0]} rows.')

    pgen = OpenAiDDPhraseGen()
    modelList = pgen.getModelNames()
    iterationCount = 1
    failedTables = tables['TableName'].tolist()
    sessionData = []

    tablesDf = tables

    while iterationCount <= maxRetries:
        for i in range(len(modelList)):
            print(f'-->\nRunning iteration {iterationCount} on AI Model `{modelList[i]}` on {len(failedTables)} tables.')

            currentSessionData, currentFailedTables = run_fname_session(pgen, dbDescr, tablesDf, glossary, phraseType, output_file, i)
            
            if len(currentFailedTables) < len(failedTables):
                sessionData = sessionData + currentSessionData
                failedTables = currentFailedTables
                tablesDf = pd.DataFrame(failedTables, columns=['TableName'])

            if len(currentFailedTables) == 0:
                break

        if len(failedTables) == 0:
            break

        iterationCount += 1    
  
    completedSessionData = addFailedTableDataToSessionData(sessionData, failedTables, glossary)
    completedSessionDf = pd.DataFrame(completedSessionData)
    sessionDf = completedSessionDf.sort_values(by=['TableName', 'ColumnName'])
    if len(failedTables) > 0:
        print(f'Unable to create Friendly Names for the following Tables: {failedTables}')

    completedNumRows = sessionDf.shape[0]
    print(f'Completed session data contains {completedNumRows} rows.')
    friendlyNames_generated = sessionDf.query("`Friendly Name`!= ''")
    friendlynameCount = friendlyNames_generated.shape[0]
    print(f'Friendly Names generated: {friendlynameCount} ({format(100 * friendlynameCount/completedNumRows, ".1f")}%).')
    
    tables.to_excel(output_file, sheet_name='Table Descriptions', index=False)
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        sessionDf.to_excel(writer, sheet_name='Column Descriptions', index=False)

    print(f'Complete generated data saved to: {output_file}')


def run_descr_session(pgen, dbDescr, tables, colDataDf, phraseType, transcriptName, model):
    current_time = datetime.now()
    sessionId = f'{transcriptName.split(".", 1)[0]}{current_time.strftime("%Y%m%d%H%M%S")}'
    print(f'Running a {phraseType} session. Session ID: {sessionId}')

    # print(f'Tables: \n{tables}')
    # print(f'Column Data: \n{colDataDf}')

    pgen.setModel(model)
    pgen.greet()

    numSuccesses = 0
    failedTables = []
    column_names = ['TableName', 'ColumnName', 'Friendly Name', 'Description']
    sessionData = pd.DataFrame(columns=column_names)

    # Iterate through each Tablename in listed in the tables data frame
    for tablesIndex, row in tables.iterrows():
        tableName = row['TableName']

        # Convert the friendly names data frame into dictionary format
        filteredColDataDf = colDataDf[colDataDf['TableName'] == tableName]
        colDat = dataframe_to_dict(filteredColDataDf)
        colDatStrList = [str(colDat)]
        numCols = filteredColDataDf.shape[0]
        print(f"   Generating {phraseType} phrases for Table Name: {tableName} with {numCols} columns...")

        try:
            responseText = pgen.sendSessionPrompt(sessionId, phraseType, dbDescr, tableName, colDatStrList)
            #Assume that response text is in standardized CSV format..

            respLength = count_csv_data_rows(responseText)
            if validate_csv_headers(sessionData, responseText):
                if respLength != numCols:
                    print(f'      !!!ERROR: Unexpected response length: {numCols} was expected but {respLength} lines were received. Check transcript for Table Name {tableName}.')
                    failedTables.append(tableName)
                else:
                    #print(f"   Success! Response received with {respLength} lines...")
                    # if '_' in responseText:
                    #     print(f'      !Warning: Underscore character detected in response.  Check transcript for Table Name {tableName}.  Cleaning up response.')
                    #     responseText = responseText.replace("\\_", " ").replace("_", " ")
                    numSuccesses += 1
                    sessionData = append_csv_to_dataframe(sessionData, responseText)
        except Exception as e:
            print(f'      !!!ERROR: Exception occured during phrase generation. Check transcript for Table Name {tableName}.')
            print(f'      Exception:{e}')
            failedTables.append(tableName)           
            if ('NETWORK ERROR DUE TO HIGH TRAFFIC' in str(e)) or ('The upstream server is timing out' in str(e)):
                print(f'   Pausing prompts for 15 seconds...')
                time.sleep(15)

        time.sleep(2)

    print(f'Session complete.  Success: {format(100* numSuccesses/len(tables), ".1f")}%')
    print(f'Failed tables: {failedTables}')
    return (sessionData, failedTables)


def count_csv_data_rows(csv_string):
    """
    Counts the number of data rows in a string containing CSV data, excluding blank or empty lines.

    :param csv_string: A string containing CSV data.
    :return: The number of non-empty data rows in the CSV data.
    :raises: ValueError if the CSV data is not properly formatted.
    """
    try:
        # Attempt to parse the CSV string using csv.reader to check for proper CSV formatting
        csv_reader = csv.reader(StringIO(csv_string))
        
        # Initialize a row count
        row_count = 0
        
        # Skip the header row
        next(csv_reader)
        
        # Iterate through the CSV data to count the rows
        for row in csv_reader:
            if any(field.strip() for field in row):  # Check if the row is not empty or doesn't contain only whitespace
                row_count += 1
        
        # If no exception was raised, return the row count
        return row_count
    except csv.Error as e:
        # If there's an issue with CSV formatting, raise an exception
        raise ValueError(f"CSV data is not properly formatted: {e}")


def validate_csv_headers(df, csv_string):
    """
    Compares the header values in the CSV string to the column names in the DataFrame.
    Throws an exception if they do not match.
    Returns True if they match.

    :param df: The Pandas DataFrame with the column names to validate against.
    :param csv_string: A string containing CSV data with a header row.
    :return: True if CSV header values match the DataFrame column names, otherwise an exception is raised.
    """
    # Convert the CSV string into a file-like object using StringIO
    csv_data_io = StringIO(csv_string)
    
    # Read the header row of the CSV data
    header = pd.read_csv(csv_data_io, nrows=0)
    
    # Extract the header columns as a list
    csv_columns = header.columns.tolist()
    
    # Extract the DataFrame columns as a list
    df_columns = df.columns.tolist()
    
    # Check if the columns match
    if csv_columns != df_columns:
        raise ValueError("CSV header values do not match DataFrame column names.")
    
    return True


def append_csv_to_dataframe(df, csv_data):
    """
    Appends CSV data to an existing DataFrame and returns the resulting DataFrame.

    :param df: The existing DataFrame to which the CSV data will be appended.
    :param csv_data: A string containing CSV data that matches the columns of the DataFrame.
    :return: A DataFrame containing the combined data.
    """
    # Check if the input DataFrame is empty
    if df.empty:
        # If empty, directly read the CSV data into the DataFrame
        new_df = pd.read_csv(StringIO(csv_data))
    else:
        # If not empty, ensure that the CSV data matches the DataFrame columns
        csv_df = pd.read_csv(StringIO(csv_data))
        if list(csv_df.columns) == list(df.columns):
            # Append the CSV data to the existing DataFrame
            new_df = pd.concat([df, csv_df], ignore_index=True)
        else:
            raise ValueError("The columns of the CSV data do not match the DataFrame.")
    
    return new_df


def run_fname_session(pgen, dbDescr, tables, glossary, phraseType, transcriptName, model):
    current_time = datetime.now()
    sessionId = f'{transcriptName.split(".", 1)[0]}{current_time.strftime("%Y%m%d%H%M%S")}'
    print(f'Running a {phraseType} session. Session ID: {sessionId}')

    pgen.setModel(model)
    pgen.greet()

    numSuccesses = 0
    failedTables = []
    sessionData = []

    # Iterate through each Tablename in listed in the tables data frame
    for tablesIndex, row in tables.iterrows():
        tableName = row['TableName']

        # Filter the Glossary DataFrame based on the Table Name
        filtered_glossary = glossary.loc[glossary['TABLENAME'] == tableName]
        
        # Extract the values of 'COLNAME' from the filtered DataFrame
        colList = filtered_glossary['COLNAME'].tolist()
        numCols = len(colList)
        print(f"   Generating {phraseType} phrases for Table Name: {tableName} with {numCols} columns...")

        try:
            responseText = pgen.sendSessionPrompt(sessionId, phraseType, dbDescr, tableName, colList)
            respLength = len(splitAndFilterLines(responseText))
            if respLength != numCols:
                print(f'      !!!ERROR: Unexpected response length: {numCols} was expected but {respLength} lines were received. Check transcript for Table Name {tableName}.')
                failedTables.append(tableName)
            else:
                #print(f"   Success! Response received with {respLength} lines...")
                if '_' in responseText:
                    print(f'      !Warning: Underscore character detected in response.  Check transcript for Table Name {tableName}.  Cleaning up response.')
                    responseText = responseText.replace("\\_", " ").replace("_", " ")
                numSuccesses += 1
                sessionData = addFriendlyNamesToSessionData(sessionData, tableName, colList, responseText)
        except Exception as e:
            print(f'      !!!ERROR: Exception occured during phrase generation. Check transcript for Table Name {tableName}.')
            print(f'      Exception:{e}')
            failedTables.append(tableName)           
            if ('NETWORK ERROR DUE TO HIGH TRAFFIC' in str(e)) or ('The upstream server is timing out' in str(e)):
                print(f'   Pausing prompts for 15 seconds...')
                time.sleep(15)

        time.sleep(2)

    print(f'Session complete.  Success: {format(100* numSuccesses/len(tables), ".1f")}%')
    print(f'Failed tables: {failedTables}')
    return (sessionData, failedTables)


def addDescriptionsToSessionData(sessionData, tableName, colData, responseText):
    updatedSessionData = sessionData
    print(f'Table: {tableName} - Response: \n{responseText}')
    # responseList = splitAndFilterLines(responseText)
    # for i in range(len(colList)):
    #     colName = colList[i]
    #     description = responseList[i]
    #     updatedSessionData.append({"TableName": tableName, "ColumnName": colName, "Friendly Name" : "zzz", "Description": description})
    return updatedSessionData


def addFriendlyNamesToSessionData(sessionData, tableName, colList, responseText):
    updatedSessionData = sessionData
    responseList = splitAndFilterLines(responseText)
    for i in range(len(colList)):
        colName = colList[i]
        friendlyName = responseList[i]
        updatedSessionData.append({"TableName": tableName, "ColumnName": colName, "Friendly Name": friendlyName})
    return updatedSessionData


def splitAndFilterLines(multiLineString):
    # Split the multi-line string into a list and filter out blank lines
    lines_list = [line for line in multiLineString.splitlines() if line.strip()]
    return lines_list


def addFailedTableDataToSessionData(sessionData, failedTables, glossary):
    updatedSessionData = sessionData

    for tableName in failedTables:
        # Filter the glossary to get the column names of the failed tables
        filtered_glosary = glossary.loc[glossary['TABLENAME'] == tableName]
        failedColList = filtered_glosary['COLNAME'].tolist()
        for colName in failedColList:
            updatedSessionData.append({"TableName": tableName, "ColumnName": colName, "Friendly Name": ""})

    return updatedSessionData


def check_input_file_exists(filename):
    if os.path.isfile(filename):
        return True
    else:
        print(f'Error: Required file {filename} does not exist.', file=sys.stderr)
        sys.exit(1)


def check_output_file(filename):
    # Check if file exists
    if not os.path.isfile(filename):
        print(f'Output file {filename} will be created.')
        return

    # Get the current date and time
    current_time = datetime.now()
    timestamp = current_time.strftime('%Y%m%d-%H%M%S')

    # Get the file path, base name, and extension
    file_path, base_name = os.path.split(filename)
    base_name, file_extension = os.path.splitext(base_name)

    # Rename the file
    new_filename = f'{base_name}_{timestamp}{file_extension}'
    new_file_path = os.path.join(file_path, new_filename)

    try:
        os.rename(filename, new_file_path)
    except PermissionError:
        print(f'Error: File {filename} is currently open in another application. Please close this file and try again.', file=sys.stderr)
        sys.exit(1)

    print(f'Existing file {filename} has been renamed to {new_filename}.')


def validate_dd_excel_file(filename):
    # Check if file exists
    if not os.path.isfile(filename):
        print(f'Error: File {filename} does not exist.', file=sys.stderr)
        sys.exit(1)

    # Check if file is an Excel file
    if not filename.lower().endswith(('.xls', '.xlsx')):
        print(f'Error: File {filename} is not an Excel file.', file=sys.stderr)
        sys.exit(1)

    # Load the workbook
    try:
        workbook = openpyxl.load_workbook(filename)
    except PermissionError:
        print(f'Error: File {filename} is currently open in another application. Please close this file and try again.', file=sys.stderr)
        sys.exit(1)

    # Check for required worksheets
    required_sheets = ['Tables', 'Glossary']
    for sheet_name in required_sheets:
        if sheet_name not in workbook.sheetnames:
            print(f'Error: File {filename} does not contain a "{sheet_name}" worksheet.', file=sys.stderr)
            sys.exit(1)

    # Check "Tables" worksheet
    table_list_sheet = workbook['Tables']
    table_list_column_names = [cell.value for cell in table_list_sheet[1]]
    if 'TableName' not in table_list_column_names:
        
        print(f'Error: "Tables" worksheet in file {filename} does not have a "TableName" column.', file=sys.stderr)
        sys.exit(1)

    # Check "Glossary" worksheet
    glossary_sheet = workbook['Glossary']
    glossary_column_names = [cell.value for cell in glossary_sheet[1]]
    if 'TABLENAME' not in glossary_column_names or 'COLNAME' not in glossary_column_names:
        print(f'Error: "Glossary" worksheet in file {filename} does not have a "TABLENAME" and "COLNAME" columns.', file=sys.stderr)
        sys.exit(1)

    # Read data from worksheets
    table_list_data = pd.read_excel(filename, sheet_name='Tables', engine='openpyxl')
    glossary_data = pd.read_excel(filename, sheet_name='Glossary', engine='openpyxl')

    return table_list_data, glossary_data


def validate_fname_excel_file(filename):
    # Check if file exists
    if not os.path.isfile(filename):
        print(f'Error: File {filename} does not exist.', file=sys.stderr)
        sys.exit(1)

    # Check if file is an Excel file
    if not filename.lower().endswith(('.xls', '.xlsx')):
        print(f'Error: File {filename} is not an Excel file.', file=sys.stderr)
        sys.exit(1)

    # Load the workbook
    try:
        workbook = openpyxl.load_workbook(filename)
    except PermissionError:
        print(f'Error: File {filename} is currently open in another application. Please close this file and try again.', file=sys.stderr)
        sys.exit(1)

    # Check for required worksheets
    required_sheets = ['Table Descriptions', 'Column Descriptions']
    for sheet_name in required_sheets:
        if sheet_name not in workbook.sheetnames:
            print(f'Error: File {filename} does not contain a "{sheet_name}" worksheet.', file=sys.stderr)
            sys.exit(1)

    # Check "Table Descriptions" worksheet
    table_list_sheet = workbook['Table Descriptions']
    table_list_column_names = [cell.value for cell in table_list_sheet[1]]
    if 'TableName' not in table_list_column_names:       
        print(f'Error: "Table Descriptions" worksheet in file {filename} does not have a "TableName" column.', file=sys.stderr)
        sys.exit(1)

    # Check "Column Descriptions" worksheet
    glossary_sheet = workbook['Column Descriptions']
    glossary_column_names = [cell.value for cell in glossary_sheet[1]]
    if 'TableName' not in glossary_column_names or 'ColumnName' not in glossary_column_names or 'Friendly Name' not in glossary_column_names:
        print(f'Error: "Column Descriptions" worksheet in file {filename} does not have a "TableName" and "ColumnName" and "Friendly Name" columns.', file=sys.stderr)
        sys.exit(1)

    # Read data from worksheets
    table_list_data = pd.read_excel(filename, sheet_name='Table Descriptions', engine='openpyxl')
    column_list_data = pd.read_excel(filename, sheet_name='Column Descriptions', engine='openpyxl')

    return table_list_data, column_list_data


###
#Command line execution starts here
###

# Initialize the argument parser
parser = argparse.ArgumentParser(description=f'Data Dictionary Phrase Generator program. Last updated: {dateLastUpdated}')

# Add arguments
parser.add_argument('-i', '--input', type=str, required=True, help='Data Dictionary Input file name')
parser.add_argument('-o', '--output', type=str, required=True, help='Output file name')
parser.add_argument('-t', '--type', type=str, required=True, help='Phrase Type (DESCRIPTION | FRIENDLYNAME)')
parser.add_argument('-d', '--descr', type=str, required=False, help='Database description (one word) - Optional')

# Parse the command line arguments
args = parser.parse_args()

# Validate the input arguments before running the program.
if args.type.upper() != pt.DESCRIPTION.name and args.type.upper() != pt.FRIENDLYNAME.name:
    print(f'Error: Value for type argument "{args.type}" is not valid.', file=sys.stderr)
    sys.exit(1)

check_input_file_exists(args.input)

check_output_file(args.output)

if args.descr is not None:
    dbDescrVal = args.descr
else:
    dbDescrVal = "relational"


# Call the appropriate "run" method with the parsed arguments

if args.type.upper() == pt.FRIENDLYNAME.name:
    tableListDf, glossaryDf = validate_dd_excel_file(args.input)
    run_friendlyName_program(dbDescrVal, tableListDf, glossaryDf, args.output, args.type.upper())
elif args.type.upper() == pt.DESCRIPTION.name:
    tableListDf, colListDf = validate_fname_excel_file(args.input)
    run_description_program(dbDescrVal, tableListDf, colListDf, args.output, args.type.upper())
else:
    sys.exit()